"""
excel.py - Excel 排列方案管理
"""

import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from core.permuter import perm_to_key


def get_excel_path(output_dir: str, date_str: str, ordered: bool,
                   selected_keys: list = None, suffix: str = None) -> str:
    """获取 Excel 路径。选了组别则带组名，如 '20260601-无序-ABD.xlsx'。"""
    mode_str = '有序' if ordered else ('无序' if suffix is None else '部分')
    parts = [date_str, mode_str]
    if selected_keys:
        parts.append(''.join(sorted(selected_keys)))
    if suffix:
        parts.append(suffix)
    return os.path.join(output_dir, f'{"-".join(parts)}.xlsx')


def load_existing_status(excel_path: str) -> dict:
    """
    读取已有 Excel，返回 {排列标识: (状态, 行号, 输出文件, 各组视频名列表)}。
    """
    result = {}
    if not os.path.exists(excel_path):
        return result
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        key_col = status_col = output_col = None
        group_cols = []  # 各组列索引

        for i, h in enumerate(headers):
            if h == '排列标识':
                key_col = i
            elif h == '拼接状态':
                status_col = i
            elif h == '输出文件':
                output_col = i
            elif h and '部分' in str(h):
                group_cols.append(i)

        if key_col is None or status_col is None:
            wb.close()
            return result

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                continue
            key = str(row[key_col]) if row[key_col] else ''
            status = str(row[status_col]) if row[status_col] else ''
            output = str(row[output_col]) if row[output_col] and output_col is not None else ''
            # 读取各组视频名
            names = [str(row[c]) for c in group_cols if c < len(row) and row[c] and str(row[c]) != '-']
            if key:
                result[key] = (status, row_num, output, names)
        wb.close()
    except Exception:
        pass
    return result


def create_excel(permutations: list, groups: dict, excel_path: str):
    """创建全新的 Excel 表格（覆盖）。"""
    group_keys = sorted(groups.keys())

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    done_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排列方案"

    headers = ['序号'] + [f'{k}部分' for k in group_keys] + \
              ['排列标识', '拼接顺序', '拼接状态', '输出文件']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, perm in enumerate(permutations):
        row = i + 2
        c = ws.cell(row=row, column=1, value=perm['index'])
        c.border = thin_border
        c.alignment = center_align

        group_video_map = {gk: vid['name'] for gk, vid in perm['videos']}
        for col_idx, key in enumerate(group_keys, 2):
            c = ws.cell(row=row, column=col_idx, value=group_video_map.get(key, '-'))
            c.border = thin_border
            c.alignment = center_align

        key_col = len(group_keys) + 2
        c = ws.cell(row=row, column=key_col, value=perm_to_key(perm))
        c.border = thin_border
        c.alignment = center_align

        concat_col = key_col + 1
        c = ws.cell(row=row, column=concat_col, value=perm['concat_str'])
        c.border = thin_border

        status_col = concat_col + 1
        c = ws.cell(row=row, column=status_col, value=perm['status'])
        c.border = thin_border
        c.alignment = center_align
        c.fill = pending_fill

        output_col = status_col + 1
        c = ws.cell(row=row, column=output_col, value=perm['output_file'])
        c.border = thin_border

    # 列宽
    ws.column_dimensions['A'].width = 8
    for i, _ in enumerate(group_keys):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 2)].width = 10
    n = len(group_keys)
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 2)].width = 18
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 3)].width = 40
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 4)].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 5)].width = 35

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)


def update_excel_status(excel_path: str, row_num: int, status: str,
                        output_file: str, groups: dict):
    """更新 Excel 某行的拼接状态。"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    group_keys = sorted(groups.keys())
    status_col = len(group_keys) + 4  # 序号 + 组列 + 标识 + 顺序 + 状态
    output_col = len(group_keys) + 5

    done_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    c = ws.cell(row=row_num, column=status_col, value=status)
    c.alignment = Alignment(horizontal='center')
    if '✅' in status:
        c.fill = done_fill
    elif '❌' in status:
        c.fill = fail_fill
    else:
        c.fill = pending_fill

    ws.cell(row=row_num, column=output_col, value=output_file)
    wb.save(excel_path)
