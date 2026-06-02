#!/usr/bin/env python3
"""
video_joiner_gui.py - 视频排列拼接工具 (GUI版)

功能：
  1. 图形界面操作，无需命令行
  2. 自动扫描输入文件夹，按首字母分组（支持 A-Z 任意组数）
  3. 有序排列（A→B→C→D）或无序排列（随机组序）
  4. 一次性生成所有排列组合，不遗漏
  5. 生成 Excel 排列方案表格，实时标记拼接状态
  6. FFmpeg 自动统一分辨率/帧率后拼接
  7. 断点续跑：已有表格时自动读取未完成的排列继续拼接

依赖：
  pip install openpyxl
  系统需安装 ffmpeg（在 PATH 中）
"""

import os
import re
import sys
import json
import random
import subprocess
import threading
from pathlib import Path
from collections import defaultdict
from itertools import product as cartesian_product

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess as sp
    sp.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], capture_output=True)
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# ═══════════════════════════════════════════════════════
# 常量 & 配置
# ═══════════════════════════════════════════════════════

VIDEO_EXTS = {'.mp4', '.mov', '.avi', '.mkv', '.flv', '.wmv', '.webm', '.m4v', '.ts'}
AUDIO_EXTS = {'.mp3', '.wav', '.aac', '.flac', '.ogg', '.m4a', '.wma'}
DEFAULT_FPS = 30
DEFAULT_CRF = 23
DEFAULT_AUDIO_BITRATE = '128k'
def _get_app_dir():
    """获取应用根目录（打包后为exe所在目录，开发时为脚本所在目录）。"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = _get_app_dir()
DEFAULT_BGM_DIR = os.path.join(APP_DIR, 'BGM')
DEFAULT_INPUT_ROOT = os.path.join(APP_DIR, 'input')
DEFAULT_OUTPUT_ROOT = os.path.join(APP_DIR, 'output')
CONFIG_PATH = os.path.join(APP_DIR, 'config.json')

# 超过此数量时弹出确认
WARN_PERM_COUNT = 500

_probe_cache = {}


# ═══════════════════════════════════════════════════════
# 配置管理
# ═══════════════════════════════════════════════════════

def load_config() -> dict:
    default = {
        'input_root': DEFAULT_INPUT_ROOT,
        'output_root': DEFAULT_OUTPUT_ROOT,
        'last_input': '',
        'last_output': '',
        'resolution': '自动 (以第一个视频为准)',
        'mode': 'ordered',
    }
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                saved = json.load(f)
            default.update(saved)
        except Exception:
            pass
    return default


def save_config(cfg: dict):
    try:
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ═══════════════════════════════════════════════════════
# 核心逻辑
# ═══════════════════════════════════════════════════════

def scan_videos(input_dir: str, rules: dict = None) -> dict:
    """
    扫描文件夹，按规则分组。

    rules 为 None 时：自动按首字母分组 (A1.mp4, B1.mp4...)
    rules 不为 None 时：按关键词匹配 ({'A': '开头', 'B': '痛点', ...})
    """
    groups = defaultdict(list)
    input_path = Path(input_dir)
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件夹不存在: {input_dir}")

    if rules:
        # 关键词匹配模式
        for f in sorted(input_path.iterdir()):
            if f.suffix.lower() not in VIDEO_EXTS:
                continue
            stem = f.stem
            matched_group = None
            for group_name, keyword in rules.items():
                if keyword in stem:
                    matched_group = group_name
                    break
            if not matched_group:
                continue
            # 提取数字：找文件名中的第一个数字序列
            num_match = re.search(r'(\d+)', stem)
            num = int(num_match.group(1)) if num_match else 0
            groups[matched_group].append({
                'file': f.name, 'name': stem, 'group': matched_group,
                'num': num, 'path': str(f.resolve())
            })
    else:
        # 自动按首字母分组 (A1.mp4, B1.mp4...)
        for f in sorted(input_path.iterdir()):
            if f.suffix.lower() not in VIDEO_EXTS:
                continue
            match = re.match(r'^([A-Za-z]+)(\d+)$', f.stem)
            if not match:
                continue
            group = match.group(1).upper()
            num = int(match.group(2))
            groups[group].append({
                'file': f.name, 'name': f.stem, 'group': group,
                'num': num, 'path': str(f.resolve())
            })

    for key in groups:
        groups[key].sort(key=lambda x: x['num'])
    return dict(sorted(groups.items()))


def probe_video(video_path: str) -> dict:
    if video_path in _probe_cache:
        return _probe_cache[video_path]
    cmd = ['ffprobe', '-v', 'quiet', '-print_format', 'json',
           '-show_streams', '-show_format', video_path]
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=30)
    except subprocess.TimeoutExpired:
        raise RuntimeError(f"ffprobe 超时: {video_path}")

    stdout = result.stdout.decode('utf-8', errors='replace') if result.stdout else ''
    if not stdout.strip():
        raise RuntimeError(f"ffprobe 无输出: {video_path}")

    data = json.loads(stdout)
    info = {'width': 1920, 'height': 1080, 'fps': DEFAULT_FPS,
            'has_audio': False, 'duration': 0.0}
    for stream in data.get('streams', []):
        if stream['codec_type'] == 'video':
            info['width'] = int(stream.get('width', 1920))
            info['height'] = int(stream.get('height', 1080))
            r_fps = stream.get('r_frame_rate', '30/1')
            try:
                num, den = map(int, r_fps.split('/'))
                info['fps'] = round(num / den) if den else DEFAULT_FPS
            except (ValueError, ZeroDivisionError):
                info['fps'] = DEFAULT_FPS
        elif stream['codec_type'] == 'audio':
            info['has_audio'] = True
    fmt = data.get('format', {})
    info['duration'] = float(fmt.get('duration', 0))
    _probe_cache[video_path] = info
    return info


# ─── 排列工具函数 ───

def perm_to_key(perm: dict) -> str:
    """排列唯一标识，如 'A1B2C3D1'。"""
    return ''.join(vid['name'] for _, vid in perm['videos'])


def perm_to_filename(perm: dict, date_str: str) -> str:
    """视频文件名，如 '20260601_开头1痛点2.mp4'。"""
    raw_key = perm_to_key(perm)
    # 清理：去掉空格，全角括号转半角，去掉特殊字符
    safe_key = raw_key.replace(' ', '').replace('\uff08', '(').replace('\uff09', ')')
    safe_key = re.sub(r'[^\w\-\(\)\u4e00-\u9fff]', '', safe_key)
    return f"{date_str}_{safe_key}.mp4"


def parse_key_to_names(key: str) -> list:
    """将 'A1B2C3D1' 解析为 ['A1', 'B2', 'C3', 'D1']。"""
    return re.findall(r'[A-Za-z]+\d+', key)


def build_video_lookup(groups: dict) -> dict:
    """构建 {video_name: video_info} 查找表。"""
    lookup = {}
    for vids in groups.values():
        for v in vids:
            lookup[v['name']] = v
    return lookup


# ─── Excel 路径 ───

def get_excel_path(output_dir: str, date_str: str, ordered: bool,
                   selected_keys: list = None, suffix: str = None) -> str:
    """获取 Excel 路径。选了组别则带组名，如 '20260601-无序-ABD.xlsx'。"""
    mode_str = '有序' if ordered else ('无序' if suffix is None else '部分')
    parts = [date_str, mode_str]
    if selected_keys:
        parts.append(''.join(sorted(selected_keys)))
    if suffix:
        parts.append(suffix)
    return os.path.join(output_dir, f'{"-".join(parts)}.xlsx')


# ─── 读取已有表格 ───

def load_existing_status(excel_path: str) -> dict:
    """
    读取已有 Excel，返回 {排列标识: (状态, 行号, 输出文件, 各组视频名列表)}。
    """
    result = {}
    if not os.path.exists(excel_path):
        return result
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        key_col = status_col = output_col = None
        group_cols = []  # 各组列索引

        for i, h in enumerate(headers):
            if h == '排列标识':
                key_col = i
            elif h == '拼接状态':
                status_col = i
            elif h == '输出文件':
                output_col = i
            elif h and '部分' in str(h):
                group_cols.append(i)

        if key_col is None or status_col is None:
            wb.close()
            return result

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                continue
            key = str(row[key_col]) if row[key_col] else ''
            status = str(row[status_col]) if row[status_col] else ''
            output = str(row[output_col]) if row[output_col] and output_col is not None else ''
            # 读取各组视频名
            names = [str(row[c]) for c in group_cols if c < len(row) and row[c] and str(row[c]) != '-']
            if key:
                result[key] = (status, row_num, output, names)
        wb.close()
    except Exception:
        pass
    return result


# ─── 生成所有排列 ───

def generate_all_permutations(groups: dict, ordered: bool = True, fixed_groups: list = None) -> list:
    """
    一次性生成所有排列组合。

    有序: 固定 A→B→C→D 顺序
    无序: 每条排列的组别顺序随机（确定性，可复现）
    部分固定: fixed_groups 中的组保持原顺序，其余组随机排列
    """
    group_keys = sorted(groups.keys())
    video_lists = [groups[k][:] for k in group_keys]  # 拷贝，避免污染原始 groups

    fixed_groups = fixed_groups or []
    fixed_set = set(fixed_groups)

    # ── 无序/部分固定时：打乱每组视频的顺序，让视频出现顺序随机 ──
    rng = random.Random(42)
    for vl in video_lists:
        rng.shuffle(vl)
    # 固定组的视频也要打乱（部分固定模式下，组内顺序仍应随机）

    permutations = []
    for i, combo in enumerate(cartesian_product(*video_lists)):
        # combo: (A组视频, B组视频, C组视频, D组视频)
        video_map = dict(zip(group_keys, combo))

        if ordered:
            order = group_keys[:]
        elif fixed_groups:
            # 部分固定模式：固定组保持原位置，其余组随机排列
            unfixed = [k for k in group_keys if k not in fixed_set]
            shuffled = unfixed[:]
            random.Random(i * 31337).shuffle(shuffled)
            order = []
            shuffled_idx = 0
            for k in group_keys:
                if k in fixed_set:
                    order.append(k)
                else:
                    order.append(shuffled[shuffled_idx])
                    shuffled_idx += 1
        else:
            # 完全无序：每个 combo 用固定种子，跨运行可复现
            order = group_keys[:]
            random.Random(i * 31337).shuffle(order)

        ordered_videos = [(k, video_map[k]) for k in order]

        permutations.append({
            'index': i + 1,
            'order': order,
            'videos': ordered_videos,
            'concat_str': ' → '.join(v['name'] for _, v in ordered_videos),
            'status': '⏳ 待拼接',
            'output_file': ''
        })

    # ── 无序/部分固定时：打乱排列顺序，避免规律性 ──
    if not ordered:
        random.Random(99).shuffle(permutations)
        # 重编序号
        for idx, perm in enumerate(permutations):
            perm['index'] = idx + 1

    return permutations


# ─── Excel 写入 ───

def create_excel(permutations: list, groups: dict, excel_path: str):
    """创建全新的 Excel 表格（覆盖）。"""
    group_keys = sorted(groups.keys())

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    done_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排列方案"

    headers = ['序号'] + [f'{k}部分' for k in group_keys] + \
              ['排列标识', '拼接顺序', '拼接状态', '输出文件']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, perm in enumerate(permutations):
        row = i + 2
        c = ws.cell(row=row, column=1, value=perm['index'])
        c.border = thin_border
        c.alignment = center_align

        group_video_map = {gk: vid['name'] for gk, vid in perm['videos']}
        for col_idx, key in enumerate(group_keys, 2):
            c = ws.cell(row=row, column=col_idx, value=group_video_map.get(key, '-'))
            c.border = thin_border
            c.alignment = center_align

        key_col = len(group_keys) + 2
        c = ws.cell(row=row, column=key_col, value=perm_to_key(perm))
        c.border = thin_border
        c.alignment = center_align

        concat_col = key_col + 1
        c = ws.cell(row=row, column=concat_col, value=perm['concat_str'])
        c.border = thin_border

        status_col = concat_col + 1
        c = ws.cell(row=row, column=status_col, value=perm['status'])
        c.border = thin_border
        c.alignment = center_align
        c.fill = pending_fill

        output_col = status_col + 1
        c = ws.cell(row=row, column=output_col, value=perm['output_file'])
        c.border = thin_border

    # 列宽
    ws.column_dimensions['A'].width = 8
    for i, _ in enumerate(group_keys):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 2)].width = 10
    n = len(group_keys)
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 2)].width = 18
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 3)].width = 40
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 4)].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(n + 5)].width = 35

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)


def update_excel_status(excel_path: str, row_num: int, status: str,
                        output_file: str, groups: dict):
    """更新 Excel 某行的拼接状态。"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    group_keys = sorted(groups.keys())
    status_col = len(group_keys) + 4  # 序号 + 组列 + 标识 + 顺序 + 状态
    output_col = len(group_keys) + 5

    done_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    c = ws.cell(row=row_num, column=status_col, value=status)
    c.alignment = Alignment(horizontal='center')
    if '✅' in status:
        c.fill = done_fill
    elif '❌' in status:
        c.fill = fail_fill
    else:
        c.fill = pending_fill

    ws.cell(row=row_num, column=output_col, value=output_file)
    wb.save(excel_path)


# ─── FFmpeg 拼接 ───

def concat_videos(video_list: list, output_path: str,
                  target_w: int = None, target_h: int = None) -> bool:
    if not video_list:
        return False
    paths = [v['path'] for v in video_list]
    n = len(paths)
    infos = [probe_video(p) for p in paths]
    if target_w is None or target_h is None:
        target_w = target_w or infos[0]['width']
        target_h = target_h or infos[0]['height']
    target_fps = infos[0]['fps']  # 以第一个视频的帧率为准
    any_has_audio = any(info['has_audio'] for info in infos)

    filter_parts = []
    for i in range(n):
        filter_parts.append(
            f"[{i}:v]scale={target_w}:{target_h}:"
            f"force_original_aspect_ratio=decrease,"
            f"pad={target_w}:{target_h}:(ow-iw)/2:(oh-ih)/2:black,"
            f"setsar=1,fps={target_fps},format=yuv420p[v{i}]"
        )
        if any_has_audio:
            if infos[i]['has_audio']:
                filter_parts.append(
                    f"[{i}:a]aformat=sample_fmts=fltp:"
                    f"sample_rates=44100:channel_layouts=stereo[a{i}]"
                )
            else:
                dur = max(infos[i]['duration'], 0.1)
                filter_parts.append(
                    f"anullsrc=r=44100:cl=stereo,atrim=0:{dur}[a{i}]"
                )

    if any_has_audio:
        concat_inputs = ''.join(f'[v{i}][a{i}]' for i in range(n))
        filter_parts.append(f"{concat_inputs}concat=n={n}:v=1:a=1[outv][outa]")
        map_args = ['-map', '[outv]', '-map', '[outa]']
    else:
        concat_inputs = ''.join(f'[v{i}]' for i in range(n))
        filter_parts.append(f"{concat_inputs}concat=n={n}:v=1:a=0[outv]")
        map_args = ['-map', '[outv]']

    filter_complex = ';'.join(filter_parts)
    cmd = ['ffmpeg', '-y', '-hide_banner', '-loglevel', 'warning']
    for p in paths:
        cmd.extend(['-i', p])
    cmd.extend(['-filter_complex', filter_complex])
    cmd.extend(map_args)
    cmd.extend(['-c:v', 'libx264', '-preset', 'medium', '-crf', str(DEFAULT_CRF),
                '-movflags', '+faststart'])
    if any_has_audio:
        cmd.extend(['-c:a', 'aac', '-b:a', DEFAULT_AUDIO_BITRATE])
    cmd.append(output_path)

    result = subprocess.run(cmd, capture_output=True, text=True)
    return result.returncode == 0


# ─── BGM 工具函数 ───

def scan_bgm_files(bgm_dir: str) -> list:
    """扫描BGM目录，返回音频文件列表。"""
    if not os.path.exists(bgm_dir):
        return []
    files = []
    for f in sorted(Path(bgm_dir).iterdir()):
        if f.suffix.lower() in AUDIO_EXTS:
            files.append({'name': f.stem, 'file': f.name, 'path': str(f.resolve())})
    return files


def add_bgm_to_video(video_path: str, bgm_path: str, output_path: str,
                     video_duration: float = None) -> bool:
    """
    给视频添加BGM。
    - BGM裁切到视频长度（尾端超出部分截断）
    - BGM循环覆盖（如果BGM比视频短）
    - 保留原视频音频，BGM混音（降低BGM音量）
    """
    if video_duration is None:
        info = probe_video(video_path)
        video_duration = info['duration']

    # 使用 amix 混合原音频和BGM，BGM循环到视频长度，然后裁切
    cmd = [
        'ffmpeg', '-y', '-hide_banner', '-loglevel', 'warning',
        '-i', video_path,
        '-stream_loop', '-1',  # BGM循环
        '-i', bgm_path,
        '-filter_complex',
        f'[1:a]atrim=0:{video_duration},asetpts=PTS-STARTPTS,afade=t=out:st={max(video_duration-2, 0)}:d=2,volume=0.3[bgm];'
        f'[0:a][bgm]amix=inputs=2:duration=first:dropout_transition=2[aout]',
        '-map', '0:v', '-map', '[aout]',
        '-c:v', 'copy',  # 视频流直接复制，不重新编码
        '-c:a', 'aac', '-b:a', DEFAULT_AUDIO_BITRATE,
        '-shortest',
        output_path
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)
    return result.returncode == 0


class BgmManager:
    """BGM分配管理器：确保BGM均匀分布。"""

    def __init__(self, bgm_files: list):
        self.bgm_files = bgm_files
        self._pool = []
        self._seed = 42

    def _refill_pool(self):
        """重新填充池：复制一份文件列表并打乱。"""
        self._pool = list(range(len(self.bgm_files)))
        random.Random(self._seed).shuffle(self._pool)
        self._seed += 1  # 每次填充用不同种子，避免重复模式

    def next(self) -> dict:
        """获取下一个BGM（均匀随机，不重复直到全部轮一遍）。"""
        if not self._pool:
            self._refill_pool()
        idx = self._pool.pop()
        return self.bgm_files[idx]


# ═══════════════════════════════════════════════════════
# GUI 应用
# ═══════════════════════════════════════════════════════

class VideoJoinerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🎬 视频排列拼接工具")
        self.root.geometry("780x680")
        self.root.resizable(True, True)
        self.root.minsize(680, 580)

        self.groups = {}
        self.group_vars = {}  # {group_name: BooleanVar}
        self.is_running = False
        self.stop_flag = False

        self.config = load_config()
        self.input_root = self.config.get('input_root', DEFAULT_INPUT_ROOT)
        self.output_root = self.config.get('output_root', DEFAULT_OUTPUT_ROOT)

        self.BG = "#f5f5f5"
        self.CARD_BG = "#ffffff"
        self.ACCENT = "#4472C4"
        self.SUCCESS = "#2e7d32"
        self.WARN = "#f57f17"
        self.FAIL = "#c62828"

        self.root.configure(bg=self.BG)

        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure('Card.TFrame', background=self.CARD_BG)
        self.style.configure('Title.TLabel', font=('Microsoft YaHei UI', 11, 'bold'),
                             background=self.CARD_BG)
        self.style.configure('Accent.TButton', font=('Microsoft YaHei UI', 10),
                             padding=(12, 6))

        self._build_ui()

    # ─── UI 构建 ───

    def _build_ui(self):
        title_frame = tk.Frame(self.root, bg=self.ACCENT, height=50)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="🎬  视频排列拼接工具",
                 font=('Microsoft YaHei UI', 16, 'bold'),
                 fg='white', bg=self.ACCENT).pack(side='left', padx=20, pady=8)
        tk.Button(title_frame, text="⚙ 设置", font=('Microsoft YaHei UI', 10),
                  fg='white', bg=self.ACCENT, relief='flat',
                  activebackground='#3a62a8', activeforeground='white',
                  cursor='hand2', padx=12, pady=2,
                  command=self._open_settings).pack(side='right', padx=20, pady=8)

        main = tk.Frame(self.root, bg=self.BG)
        main.pack(fill='both', expand=True, padx=16, pady=10)

        self._build_folder_card(main)
        self._build_params_card(main)
        self._build_preview_card(main)
        self._build_action_card(main)
        self._build_log_area(main)

    def _build_folder_card(self, parent):
        card = ttk.Frame(parent, style='Card.TFrame')
        card.pack(fill='x', pady=(0, 8))
        inner = tk.Frame(card, bg=self.CARD_BG)
        inner.pack(fill='x', padx=16, pady=12)

        ttk.Label(inner, text="📁 文件夹设置", style='Title.TLabel').pack(anchor='w')

        row1 = tk.Frame(inner, bg=self.CARD_BG)
        row1.pack(fill='x', pady=(8, 4))
        tk.Label(row1, text="素材目录：", font=('Microsoft YaHei UI', 10),
                 bg=self.CARD_BG, width=10, anchor='e').pack(side='left')
        self.input_var = tk.StringVar(value=self.config.get('last_input', ''))
        tk.Entry(row1, textvariable=self.input_var, font=('Consolas', 10),
                 state='readonly', readonlybackground='#fafafa').pack(
            side='left', fill='x', expand=True, padx=(0, 8))
        ttk.Button(row1, text="浏览...", style='Accent.TButton',
                   command=self._browse_input).pack(side='right')

        tk.Label(inner, text=f"素材根目录: {self.input_root}",
                 font=('Microsoft YaHei UI', 8), bg=self.CARD_BG, fg='#999').pack(
            anchor='w', padx=(78, 0))

        row2 = tk.Frame(inner, bg=self.CARD_BG)
        row2.pack(fill='x', pady=4)
        tk.Label(row2, text="输出目录：", font=('Microsoft YaHei UI', 10),
                 bg=self.CARD_BG, width=10, anchor='e').pack(side='left')
        self.output_var = tk.StringVar(value=self.config.get('last_output', ''))
        tk.Entry(row2, textvariable=self.output_var, font=('Consolas', 10),
                 state='readonly', readonlybackground='#fafafa').pack(
            side='left', fill='x', expand=True, padx=(0, 8))
        ttk.Button(row2, text="浏览...", style='Accent.TButton',
                   command=self._browse_output).pack(side='right')

        tk.Label(inner, text=f"输出根目录: {self.output_root}（自动按日期分文件夹）",
                 font=('Microsoft YaHei UI', 8), bg=self.CARD_BG, fg='#999').pack(
            anchor='w', padx=(78, 0))

        # ── 组规则 ──
        sep = ttk.Separator(inner, orient='horizontal')
        sep.pack(fill='x', pady=(10, 6))

        rule_top = tk.Frame(inner, bg=self.CARD_BG)
        rule_top.pack(fill='x')
        tk.Label(rule_top, text="组规则：", font=('Microsoft YaHei UI', 10, 'bold'),
                 bg=self.CARD_BG).pack(side='left')
        tk.Label(rule_top, text="（留空则自动按首字母分组 A1.mp4 B1.mp4...）",
                 font=('Microsoft YaHei UI', 8), bg=self.CARD_BG, fg='#999').pack(
            side='left', padx=(4, 0))

        self.rules_var = tk.StringVar(value=self.config.get('rules', ''))
        rules_entry = tk.Entry(inner, textvariable=self.rules_var,
                               font=('Microsoft YaHei UI', 10))
        rules_entry.pack(fill='x', pady=(4, 0))

        tk.Label(inner, text="格式: A=开头,B=痛点,C=结尾,D=效果  （组名=关键词，逗号分隔）",
                 font=('Microsoft YaHei UI', 8), bg=self.CARD_BG, fg='#666').pack(
            anchor='w')
        tk.Label(inner, text="文件名包含关键词即归入该组，如 '开头_产品介绍.mp4' → A组",
                 font=('Microsoft YaHei UI', 8), bg=self.CARD_BG, fg='#999').pack(
            anchor='w')

    def _build_params_card(self, parent):
        card = ttk.Frame(parent, style='Card.TFrame')
        card.pack(fill='x', pady=(0, 8))
        inner = tk.Frame(card, bg=self.CARD_BG)
        inner.pack(fill='x', padx=16, pady=12)

        ttk.Label(inner, text="⚙️ 参数设置", style='Title.TLabel').pack(anchor='w')

        row = tk.Frame(inner, bg=self.CARD_BG)
        row.pack(fill='x', pady=(8, 0))

        f1 = tk.LabelFrame(row, text="排列模式", font=('Microsoft YaHei UI', 9),
                           bg=self.CARD_BG, padx=10, pady=6)
        f1.pack(side='left', padx=(0, 16))
        self.mode_var = tk.StringVar(value=self.config.get('mode', 'ordered'))
        tk.Radiobutton(f1, text="有序 (A→B→C→D)", variable=self.mode_var,
                       value='ordered', bg=self.CARD_BG,
                       font=('Microsoft YaHei UI', 9), command=self._on_mode_change).pack(anchor='w')
        tk.Radiobutton(f1, text="无序 (随机组序)", variable=self.mode_var,
                       value='unordered', bg=self.CARD_BG,
                       font=('Microsoft YaHei UI', 9), command=self._on_mode_change).pack(anchor='w')
        tk.Radiobutton(f1, text="部分固定 (自定义)", variable=self.mode_var,
                       value='partial', bg=self.CARD_BG,
                       font=('Microsoft YaHei UI', 9), command=self._on_mode_change).pack(anchor='w')

        f2 = tk.LabelFrame(row, text="拼接数量", font=('Microsoft YaHei UI', 9),
                           bg=self.CARD_BG, padx=10, pady=6)
        f2.pack(side='left', padx=(0, 24))
        self.count_var = tk.StringVar(value=str(self.config.get('count', 100)))
        count_spin = tk.Spinbox(f2, from_=1, to=99999, width=8,
                                textvariable=self.count_var,
                                font=('Microsoft YaHei UI', 10),
                                buttonbackground='#ddd')
        count_spin.pack()
        self.count_hint = tk.Label(f2, text="（表格自动生成全部，此处控制拼接视频数）",
                                   font=('Microsoft YaHei UI', 8),
                                   bg=self.CARD_BG, fg='#888')
        self.count_hint.pack()

        # ── 部分固定设置 ──
        self.partial_frame = tk.LabelFrame(row, text="固定组别设置", font=('Microsoft YaHei UI', 9),
                                           bg=self.CARD_BG, padx=10, pady=6)
        self.partial_frame.pack(side='left', padx=(0, 16))
        self.partial_frame.pack_forget()  # 默认隐藏

        self.partial_fixed_vars = {}  # 各组是否固定的复选框变量
        self.partial_fixed_frame = tk.Frame(self.partial_frame, bg=self.CARD_BG)
        self.partial_fixed_frame.pack()
        tk.Label(self.partial_fixed_frame, text="选择要固定的组：", font=('Microsoft YaHei UI', 9),
                 bg=self.CARD_BG, fg='#666').pack(side='left')

        f3 = tk.LabelFrame(row, text="输出分辨率", font=('Microsoft YaHei UI', 9),
                           bg=self.CARD_BG, padx=10, pady=6)
        f3.pack(side='left', padx=(0, 16))
        self.res_var = tk.StringVar(
            value=self.config.get('resolution', '自动 (以第一个视频为准)'))
        ttk.Combobox(f3, textvariable=self.res_var,
                     values=['自动 (以第一个视频为准)', '1920x1080', '1280x720',
                             '854x480', '3840x2160'],
                     state='readonly', width=24,
                     font=('Microsoft YaHei UI', 9)).pack()

        # ── BGM 设置 ──
        f4 = tk.LabelFrame(row, text="背景音乐", font=('Microsoft YaHei UI', 9),
                           bg=self.CARD_BG, padx=10, pady=6)
        f4.pack(side='left')
        self.use_bgm_var = tk.BooleanVar(value=self.config.get('use_bgm', False))
        tk.Checkbutton(f4, text="添加BGM", variable=self.use_bgm_var,
                       font=('Microsoft YaHei UI', 9), bg=self.CARD_BG,
                       command=self._on_bgm_toggle).pack(anchor='w')
        self.bgm_info_label = tk.Label(f4, text="未选择", font=('Microsoft YaHei UI', 8),
                                       bg=self.CARD_BG, fg='#999')
        self.bgm_info_label.pack(anchor='w')
        self.bgm_select_btn = tk.Button(f4, text="选择BGM...", font=('Microsoft YaHei UI', 8),
                                        relief='flat', bg='#e8eaf6', padx=6, cursor='hand2',
                                        command=self._open_bgm_picker)
        self.bgm_select_btn.pack(anchor='w', pady=(2, 0))

        # BGM状态
        self.selected_bgm_files = []  # 用户选中的BGM文件列表
        self.bgm_dir = self.config.get('bgm_dir', DEFAULT_BGM_DIR)

    def _build_preview_card(self, parent):
        card = ttk.Frame(parent, style='Card.TFrame')
        card.pack(fill='x', pady=(0, 8))
        inner = tk.Frame(card, bg=self.CARD_BG)
        inner.pack(fill='x', padx=16, pady=12)

        top_row = tk.Frame(inner, bg=self.CARD_BG)
        top_row.pack(fill='x')
        ttk.Label(top_row, text="📋 视频分组预览", style='Title.TLabel').pack(side='left')
        ttk.Button(top_row, text="🔄 重新扫描", style='Accent.TButton',
                   command=self._scan_folder).pack(side='right')

        self.preview_frame = tk.Frame(inner, bg=self.CARD_BG)
        self.preview_frame.pack(fill='x', pady=(8, 0))

        tk.Label(self.preview_frame, text="请先选择素材目录",
                 font=('Microsoft YaHei UI', 10),
                 bg=self.CARD_BG, fg='#999').pack()

    def _build_action_card(self, parent):
        card = tk.Frame(parent, bg=self.CARD_BG, relief='flat',
                        highlightthickness=1, highlightbackground='#ddd')
        card.pack(fill='x', pady=(0, 8))
        inner = tk.Frame(card, bg=self.CARD_BG)
        inner.pack(fill='x', padx=16, pady=12)

        btn_row = tk.Frame(inner, bg=self.CARD_BG)
        btn_row.pack(fill='x')

        self.start_btn = tk.Button(
            btn_row, text="▶  开始拼接", font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.ACCENT, fg='white', activebackground='#3a62a8',
            activeforeground='white', relief='flat', padx=30, pady=8,
            cursor='hand2', command=self._start_process)
        self.start_btn.pack(side='left', padx=(0, 12))

        self.excel_btn = tk.Button(
            btn_row, text="📊 仅生成表格", font=('Microsoft YaHei UI', 10),
            bg='#e8eaf6', fg='#333', activebackground='#c5cae9',
            relief='flat', padx=16, pady=8, cursor='hand2',
            command=self._generate_excel_only)
        self.excel_btn.pack(side='left', padx=(0, 12))

        self.stop_btn = tk.Button(
            btn_row, text="⏹ 停止", font=('Microsoft YaHei UI', 10),
            bg='#ffcdd2', fg='#c62828', activebackground='#ef9a9a',
            relief='flat', padx=16, pady=8, cursor='hand2',
            state='disabled', command=self._stop_process)
        self.stop_btn.pack(side='left')

        prog_frame = tk.Frame(inner, bg=self.CARD_BG)
        prog_frame.pack(fill='x', pady=(10, 0))

        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(prog_frame, variable=self.progress_var,
                        maximum=100, length=400).pack(
            fill='x', side='left', expand=True, padx=(0, 10))

        self.progress_label = tk.Label(prog_frame, text="0/0 (0%)",
                                       font=('Consolas', 10), bg=self.CARD_BG, fg='#555')
        self.progress_label.pack(side='right')

        self.result_label = tk.Label(inner, text="", font=('Microsoft YaHei UI', 10),
                                     bg=self.CARD_BG)
        self.result_label.pack(anchor='w', pady=(6, 0))

    def _build_log_area(self, parent):
        log_frame = tk.Frame(parent, bg=self.BG)
        log_frame.pack(fill='both', expand=True)

        tk.Label(log_frame, text="📝 运行日志", font=('Microsoft YaHei UI', 9, 'bold'),
                 bg=self.BG, fg='#666').pack(anchor='w')

        self.log_text = tk.Text(log_frame, height=8, font=('Consolas', 9),
                                bg='#1e1e1e', fg='#d4d4d4',
                                insertbackground='white', relief='flat',
                                wrap='word', state='disabled')
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        self.log_text.pack(fill='both', expand=True, pady=(4, 0))

        self.log_text.tag_configure('info', foreground='#d4d4d4')
        self.log_text.tag_configure('success', foreground='#6a9955')
        self.log_text.tag_configure('warn', foreground='#dcdcaa')
        self.log_text.tag_configure('error', foreground='#f44747')
        self.log_text.tag_configure('title', foreground='#569cd6',
                                    font=('Consolas', 9, 'bold'))

    # ─── 日志 ───

    def _log(self, msg, tag='info'):
        self.log_text.configure(state='normal')
        self.log_text.insert('end', msg + '\n', tag)
        self.log_text.see('end')
        self.log_text.configure(state='disabled')

    def _log_clear(self):
        self.log_text.configure(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.configure(state='disabled')

    # ─── 设置 ───

    def _open_settings(self):
        win = tk.Toplevel(self.root)
        win.title("⚙ 设置")
        win.geometry("560x260")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()
        win.configure(bg=self.CARD_BG)

        win.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 560) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 220) // 2
        win.geometry(f"560x260+{x}+{y}")

        tk.Label(win, text="⚙  路径设置", font=('Microsoft YaHei UI', 14, 'bold'),
                 bg=self.CARD_BG, fg=self.ACCENT).pack(pady=(16, 12))

        row1 = tk.Frame(win, bg=self.CARD_BG)
        row1.pack(fill='x', padx=24, pady=6)
        tk.Label(row1, text="素材根目录：", font=('Microsoft YaHei UI', 10),
                 bg=self.CARD_BG, width=12, anchor='e').pack(side='left')
        input_var = tk.StringVar(value=self.input_root)
        tk.Entry(row1, textvariable=input_var, font=('Consolas', 10),
                 width=42).pack(side='left', padx=(0, 8))
        ttk.Button(row1, text="浏览",
                   command=lambda: self._settings_browse(input_var)).pack(side='right')

        row2 = tk.Frame(win, bg=self.CARD_BG)
        row2.pack(fill='x', padx=24, pady=6)
        tk.Label(row2, text="输出根目录：", font=('Microsoft YaHei UI', 10),
                 bg=self.CARD_BG, width=12, anchor='e').pack(side='left')
        output_var = tk.StringVar(value=self.output_root)
        tk.Entry(row2, textvariable=output_var, font=('Consolas', 10),
                 width=42).pack(side='left', padx=(0, 8))
        ttk.Button(row2, text="浏览",
                   command=lambda: self._settings_browse(output_var)).pack(side='right')

        row3 = tk.Frame(win, bg=self.CARD_BG)
        row3.pack(fill='x', padx=24, pady=6)
        tk.Label(row3, text="BGM 目录：", font=('Microsoft YaHei UI', 10),
                 bg=self.CARD_BG, width=12, anchor='e').pack(side='left')
        bgm_var = tk.StringVar(value=self.bgm_dir)
        tk.Entry(row3, textvariable=bgm_var, font=('Consolas', 10),
                 width=42).pack(side='left', padx=(0, 8))
        ttk.Button(row3, text="浏览",
                   command=lambda: self._settings_browse(bgm_var)).pack(side='right')

        btn_frame = tk.Frame(win, bg=self.CARD_BG)
        btn_frame.pack(fill='x', padx=24, pady=(12, 0))

        def save_and_close():
            self.input_root = input_var.get()
            self.output_root = output_var.get()
            self.bgm_dir = bgm_var.get()
            self.config['input_root'] = self.input_root
            self.config['output_root'] = self.output_root
            self.config['bgm_dir'] = self.bgm_dir
            save_config(self.config)
            win.destroy()
            self._log(f"设置已保存: 素材={self.input_root}, 输出={self.output_root}, BGM={self.bgm_dir}", 'title')

        tk.Button(btn_frame, text="保存", font=('Microsoft YaHei UI', 10, 'bold'),
                  bg=self.ACCENT, fg='white', relief='flat', padx=24, pady=4,
                  cursor='hand2', command=save_and_close).pack(side='right')
        tk.Button(btn_frame, text="取消", font=('Microsoft YaHei UI', 10),
                  bg='#e0e0e0', fg='#333', relief='flat', padx=16, pady=4,
                  cursor='hand2', command=win.destroy).pack(side='right', padx=(0, 8))

    def _settings_browse(self, var):
        path = filedialog.askdirectory(title="选择目录")
        if path:
            var.set(path)

    # ─── BGM ───

    def _on_bgm_toggle(self):
        """BGM开关切换。"""
        if self.use_bgm_var.get():
            # 打开时自动扫描BGM目录
            if not self.selected_bgm_files:
                all_bgm = scan_bgm_files(self.bgm_dir)
                if all_bgm:
                    self.selected_bgm_files = all_bgm
                    self.bgm_info_label.config(
                        text=f"已加载 {len(all_bgm)} 首（目录）", fg='#2e7d32')
                else:
                    self.bgm_info_label.config(text="⚠️ BGM目录为空", fg='#e65100')
        else:
            self.bgm_info_label.config(text="未选择", fg='#999')
        self.config['use_bgm'] = self.use_bgm_var.get()
        save_config(self.config)

    def _open_bgm_picker(self):
        """打开BGM选择弹窗。"""
        all_bgm = scan_bgm_files(self.bgm_dir)
        if not all_bgm:
            messagebox.showwarning("提示", f"BGM目录为空或不存在：\n{self.bgm_dir}")
            return

        win = tk.Toplevel(self.root)
        win.title("🎵 选择背景音乐")
        win.geometry("500x420")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()
        win.configure(bg=self.CARD_BG)

        win.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 500) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 420) // 2
        win.geometry(f"500x420+{x}+{y}")

        tk.Label(win, text="🎵  选择背景音乐", font=('Microsoft YaHei UI', 14, 'bold'),
                 bg=self.CARD_BG, fg=self.ACCENT).pack(pady=(16, 8))
        tk.Label(win, text=f"目录: {self.bgm_dir}", font=('Microsoft YaHei UI', 8),
                 bg=self.CARD_BG, fg='#999').pack()

        # 全选/全不选
        btn_row = tk.Frame(win, bg=self.CARD_BG)
        btn_row.pack(fill='x', padx=24, pady=(12, 4))

        bgm_vars = []  # [(BooleanVar, bgm_dict)]

        def select_all():
            for var, _ in bgm_vars:
                var.set(True)

        def deselect_all():
            for var, _ in bgm_vars:
                var.set(False)

        tk.Button(btn_row, text="全选", font=('Microsoft YaHei UI', 8),
                  relief='flat', bg='#e8eaf6', padx=8, cursor='hand2',
                  command=select_all).pack(side='left')
        tk.Button(btn_row, text="全不选", font=('Microsoft YaHei UI', 8),
                  relief='flat', bg='#e8eaf6', padx=8, cursor='hand2',
                  command=deselect_all).pack(side='left', padx=(8, 0))
        tk.Label(btn_row, text=f"共 {len(all_bgm)} 首",
                 font=('Microsoft YaHei UI', 8), bg=self.CARD_BG, fg='#666').pack(side='right')

        # 列表区
        list_frame = tk.Frame(win, bg=self.CARD_BG)
        list_frame.pack(fill='both', expand=True, padx=24, pady=(4, 8))

        canvas = tk.Canvas(list_frame, bg=self.CARD_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=canvas.yview)
        scrollable = tk.Frame(canvas, bg=self.CARD_BG)

        scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scrollable, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        # 默认选中状态：基于之前的选择
        selected_names = {f['name'] for f in self.selected_bgm_files}

        for bgm in all_bgm:
            var = tk.BooleanVar(value=(bgm['name'] in selected_names or len(selected_names) == 0))
            bgm_vars.append((var, bgm))
            tk.Checkbutton(scrollable, text=bgm['name'], variable=var,
                           font=('Microsoft YaHei UI', 9), bg=self.CARD_BG,
                           anchor='w').pack(fill='x', padx=8)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 确认按钮
        def confirm():
            self.selected_bgm_files = [bgm for var, bgm in bgm_vars if var.get()]
            count = len(self.selected_bgm_files)
            if count > 0:
                self.bgm_info_label.config(
                    text=f"已选 {count} 首", fg='#2e7d32')
                self.use_bgm_var.set(True)
            else:
                self.bgm_info_label.config(text="未选择", fg='#999')
                self.use_bgm_var.set(False)
            self.config['use_bgm'] = self.use_bgm_var.get()
            save_config(self.config)
            win.destroy()

        tk.Button(win, text="确认选择", font=('Microsoft YaHei UI', 10, 'bold'),
                  bg=self.ACCENT, fg='white', relief='flat', padx=24, pady=6,
                  cursor='hand2', command=confirm).pack(pady=(0, 16))

    # ─── 文件夹浏览 ───

    def _browse_input(self):
        path = filedialog.askdirectory(title="选择素材文件夹", initialdir=self.input_root)
        if path:
            self.input_var.set(path)
            folder_name = os.path.basename(path)
            self.output_var.set(os.path.join(self.output_root, folder_name))
            self._scan_folder()

    def _browse_output(self):
        path = filedialog.askdirectory(title="选择输出文件夹", initialdir=self.output_root)
        if path:
            self.output_var.set(path)

    # ─── 规则解析 ───

    def _parse_rules(self) -> dict:
        """解析组规则文本，返回 {组名: 关键词}。"""
        text = self.rules_var.get().strip()
        if not text:
            return {}
        rules = {}
        # 支持半角逗号和全角逗号
        text = text.replace('，', ',')
        for part in text.split(','):
            part = part.strip()
            if '=' in part:
                name, keyword = part.split('=', 1)
                name = name.strip().upper()
                keyword = keyword.strip()
                if name and keyword:
                    rules[name] = keyword
        return rules

    # ─── 扫描 ───

    def _scan_folder(self):
        input_dir = self.input_var.get()
        if not input_dir:
            messagebox.showwarning("提示", "请先选择素材目录！")
            return

        for widget in self.preview_frame.winfo_children():
            widget.destroy()

        try:
            rules = self._parse_rules()
            self.groups = scan_videos(input_dir, rules=rules if rules else None)
        except FileNotFoundError:
            messagebox.showerror("错误", "素材目录不存在！")
            return

        if not self.groups:
            rules = self._parse_rules()
            hint = "未找到包含关键词的视频" if rules else "未找到符合命名规则的视频 (A1.mp4, B1.mp4, ...)"
            tk.Label(self.preview_frame, text=f"⚠️ {hint}",
                     font=('Microsoft YaHei UI', 10),
                     bg=self.CARD_BG, fg='#e65100').pack()
            return

        group_keys = sorted(self.groups.keys())
        total_videos = sum(len(v) for v in self.groups.values())
        max_perms = 1
        for v in self.groups.values():
            max_perms *= len(v)

        # 分组卡片
        cards_frame = tk.Frame(self.preview_frame, bg=self.CARD_BG)
        cards_frame.pack(fill='x')
        colors = ['#1565c0', '#2e7d32', '#e65100', '#6a1b9a', '#00695c',
                  '#c62828', '#37474f', '#ad1457']
        for i, key in enumerate(group_keys):
            videos = self.groups[key]
            color = colors[i % len(colors)]
            cf = tk.Frame(cards_frame, bg=color, padx=10, pady=6)
            cf.pack(side='left', padx=(0, 6), pady=2)
            tk.Label(cf, text=f" {key} 组 ",
                     font=('Microsoft YaHei UI', 10, 'bold'),
                     fg='white', bg=color).pack()
            names = ', '.join(v['name'] for v in videos)
            tk.Label(cf, text=names, font=('Microsoft YaHei UI', 8),
                     fg='#e0e0e0', bg=color).pack()

        # ── 组别选择复选框 ──
        select_frame = tk.Frame(self.preview_frame, bg=self.CARD_BG)
        select_frame.pack(fill='x', pady=(8, 0))
        tk.Label(select_frame, text="选择参与排列的组别：",
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 bg=self.CARD_BG, fg='#333').pack(side='left')

        self.group_vars = {}
        for i, key in enumerate(group_keys):
            var = tk.BooleanVar(value=True)
            self.group_vars[key] = var
            color = colors[i % len(colors)]
            cb = tk.Checkbutton(select_frame, text=f"{key}组", variable=var,
                                font=('Microsoft YaHei UI', 10, 'bold'),
                                bg=self.CARD_BG, fg=color, selectcolor='#e8eaf6',
                                activebackground=self.CARD_BG,
                                command=self._on_group_selection_change)
            cb.pack(side='left', padx=(12, 0))
            # 同时为部分固定模式创建/保留复选框变量
            if key not in self.partial_fixed_vars:
                self.partial_fixed_vars[key] = tk.BooleanVar(value=self.config.get(f'fixed_{key}', False))

        # 清理已不存在的组的旧变量
        stale_keys = [k for k in self.partial_fixed_vars if k not in group_keys]
        for k in stale_keys:
            del self.partial_fixed_vars[k]

        # 全选/全不选
        btn_frame = tk.Frame(select_frame, bg=self.CARD_BG)
        btn_frame.pack(side='left', padx=(16, 0))
        tk.Button(btn_frame, text="全选", font=('Microsoft YaHei UI', 8),
                  relief='flat', bg='#e8eaf6', padx=6, cursor='hand2',
                  command=lambda: self._toggle_all_groups(True)).pack(side='left')
        tk.Button(btn_frame, text="全不选", font=('Microsoft YaHei UI', 8),
                  relief='flat', bg='#e8eaf6', padx=6, cursor='hand2',
                  command=lambda: self._toggle_all_groups(False)).pack(side='left', padx=(4, 0))

        # 排列数量显示
        self.perm_count_label = tk.Label(self.preview_frame, text="",
                                         font=('Microsoft YaHei UI', 10, 'bold'),
                                         bg=self.CARD_BG, fg='#333')
        self.perm_count_label.pack(anchor='w', pady=(6, 0))

        self._update_perm_count()

        # 显示已有表格信息
        self._show_existing_info(input_dir)

        # 如果当前是部分固定模式，刷新固定组UI
        if self.mode_var.get() == 'partial':
            self._update_partial_fixed_ui()

        self._log(f"扫描完成: {len(group_keys)} 组, {total_videos} 个视频, "
                  f"排列 {max_perms:,} 种", 'title')

    def _on_group_selection_change(self):
        """组别选择变化时：更新排列数量 + 刷新固定组UI + 清理取消勾选组的固定状态。"""
        selected_keys = set(self._get_selected_groups().keys())
        # 清理已取消勾选组的固定状态
        for key, var in self.partial_fixed_vars.items():
            if key not in selected_keys and var.get():
                var.set(False)
        self._update_perm_count()
        # 如果是部分固定模式，同步刷新固定组UI
        if self.mode_var.get() == 'partial':
            self._update_partial_fixed_ui()

    def _toggle_all_groups(self, value: bool):
        """全选/全不选组别。"""
        for var in self.group_vars.values():
            var.set(value)
        self._on_group_selection_change()

    def _on_mode_change(self):
        """排列模式切换时更新UI。"""
        if self.mode_var.get() == 'partial':
            self.partial_frame.pack(side='left', padx=(0, 16))
            self._update_partial_fixed_ui()
        else:
            self.partial_frame.pack_forget()

    def _update_partial_fixed_ui(self):
        """更新部分固定组的复选框UI（只显示选中的组）。"""
        # 清空现有
        for widget in self.partial_fixed_frame.winfo_children():
            widget.destroy()

        tk.Label(self.partial_fixed_frame, text="固定组：", font=('Microsoft YaHei UI', 9),
                 bg=self.CARD_BG, fg='#666').pack(side='left')

        if not self.groups:
            return

        selected_keys = set(self._get_selected_groups().keys())
        colors = ['#1565c0', '#2e7d32', '#e65100', '#6a1b9a', '#00695c',
                  '#c62828', '#37474f', '#ad1457']

        for i, key in enumerate(sorted(self.groups.keys())):
            if key not in selected_keys:
                continue  # 只显示选中的组
            if key not in self.partial_fixed_vars:
                self.partial_fixed_vars[key] = tk.BooleanVar(value=False)
            color = colors[i % len(colors)]
            cb = tk.Checkbutton(self.partial_fixed_frame, text=key, variable=self.partial_fixed_vars[key],
                                font=('Microsoft YaHei UI', 9, 'bold'),
                                bg=self.CARD_BG, fg=color, selectcolor='#e8eaf6',
                                activebackground=self.CARD_BG,
                                command=self._update_perm_count)
            cb.pack(side='left', padx=(8, 0))

    def _get_selected_groups(self) -> dict:
        """返回用户选中的组别。"""
        return {k: v for k, v in self.groups.items()
                if k in self.group_vars and self.group_vars[k].get()}

    def _update_perm_count(self):
        """更新排列数量显示。"""
        selected = self._get_selected_groups()
        if not selected:
            self.perm_count_label.config(text="⚠️ 请至少选择一组", fg='#e65100')
            return
        max_perms = 1
        for v in selected.values():
            max_perms *= len(v)
        if max_perms < 2:
            self.perm_count_label.config(
                text=f"选中 {len(selected)} 组 · 排列数: {max_perms}（至少需要2组）",
                fg='#e65100')
        else:
            self.perm_count_label.config(
                text=f"选中 {len(selected)} 组 · 所有排列共 {max_perms:,} 种",
                fg='#333')

    def _show_existing_info(self, input_dir: str):
        folder_name = os.path.basename(input_dir)
        output_dir = os.path.join(self.output_root, folder_name)
        info_parts = []
        # 扫描所有已有表格
        if os.path.exists(output_dir):
            for f in sorted(Path(output_dir).glob('*.xlsx')):
                status = load_existing_status(str(f))
                if status:
                    done = sum(1 for v in status.values() if '✅' in v[0])
                    info_parts.append(f"{f.stem}: {len(status)}条 ({done}✅)")

        if info_parts:
            tk.Label(self.preview_frame,
                     text="📊 已有表格: " + " | ".join(info_parts),
                     font=('Microsoft YaHei UI', 10),
                     bg=self.CARD_BG, fg='#1565c0').pack(anchor='w', pady=(4, 0))

    # ─── 参数解析 ───

    def _get_resolution(self):
        res_str = self.res_var.get()
        if '自动' in res_str:
            return None, None
        try:
            return tuple(map(int, res_str.split('x')))
        except ValueError:
            return None, None

    def _get_params(self):
        input_dir = self.input_var.get()
        output_dir = self.output_var.get()

        if not input_dir:
            messagebox.showwarning("提示", "请选择素材目录！")
            return None
        if not self.groups:
            messagebox.showwarning("提示", "未扫描到视频，请检查素材目录！")
            return None

        # 获取用户选中的组别
        selected_groups = self._get_selected_groups()
        if not selected_groups:
            messagebox.showwarning("提示", "请至少选择一组视频！")
            return None
        if len(selected_groups) < 2:
            messagebox.showwarning("提示", "至少需要选择两组才能排列！")
            return None

        if not output_dir:
            folder_name = os.path.basename(input_dir)
            output_dir = os.path.join(self.output_root, folder_name)
            self.output_var.set(output_dir)

        try:
            count = int(self.count_var.get())
        except ValueError:
            messagebox.showwarning("提示", "拼接数量必须是数字！")
            return None
        if count < 1:
            messagebox.showwarning("提示", "拼接数量必须 >= 1！")
            return None

        mode = self.mode_var.get()
        ordered = (mode == 'ordered')
        # 获取固定组设置
        fixed_groups = []
        if mode == 'partial':
            fixed_groups = sorted([k for k, v in self.partial_fixed_vars.items() if v.get()])
            if not fixed_groups:
                messagebox.showwarning("提示", "部分固定模式下请至少选择一组固定！")
                return None
            if len(fixed_groups) >= len(selected_groups):
                messagebox.showwarning("提示", "固定组不能等于或超过总组数！\n至少留一组可变。")
                return None

        res_w, res_h = self._get_resolution()

        os.makedirs(output_dir, exist_ok=True)

        # BGM设置
        use_bgm = self.use_bgm_var.get()
        bgm_files = self.selected_bgm_files if use_bgm else []
        if use_bgm and not bgm_files:
            # 尝试自动加载
            bgm_files = scan_bgm_files(self.bgm_dir)
            if not bgm_files:
                messagebox.showwarning("提示", "已勾选BGM但未选择音乐文件！")
                return None

        # 保存配置
        self.config['last_input'] = input_dir
        self.config['last_output'] = output_dir
        self.config['mode'] = self.mode_var.get()
        self.config['count'] = count
        self.config['resolution'] = self.res_var.get()
        self.config['rules'] = self.rules_var.get()
        self.config['use_bgm'] = use_bgm
        # 保存部分固定的设置
        for key, var in self.partial_fixed_vars.items():
            self.config[f'fixed_{key}'] = var.get()
        save_config(self.config)

        return selected_groups, ordered, fixed_groups, count, res_w, res_h, output_dir, bgm_files

    # ─── 仅生成表格 ───

    def _generate_excel_only(self):
        params = self._get_params()
        if not params:
            return

        groups, ordered, fixed_groups, _, _, _, output_dir, _ = params
        date_str = os.path.basename(self.input_var.get())
        selected_keys = list(groups.keys())
        # 部分固定模式下，表格命名加上固定组信息
        excel_suffix = f"固定{''.join(fixed_groups)}" if fixed_groups else None
        excel_path = get_excel_path(output_dir, date_str, ordered, selected_keys, excel_suffix)

        # 检查已有表格
        existing = load_existing_status(excel_path)
        if existing:
            done_count = sum(1 for v in existing.values() if '✅' in v[0])
            answer = messagebox.askyesno(
                "表格已存在",
                f"已存在表格: {os.path.basename(excel_path)}\n"
                f"共 {len(existing)} 条排列 ({done_count} 条已完成)\n\n"
                f"重新生成会覆盖现有表格，确认？")
            if not answer:
                return

        self._log_clear()
        self._log("📊 仅生成表格模式", 'title')

        perms = generate_all_permutations(groups, ordered, fixed_groups)

        # 高数量确认
        if len(perms) > WARN_PERM_COUNT:
            answer = messagebox.askyesno(
                "确认",
                f"将生成 {len(perms):,} 种排列组合，数量较大。\n\n确认继续？")
            if not answer:
                return

        create_excel(perms, groups, excel_path)

        self._log(f"✅ 表格已生成: {excel_path}", 'success')
        self._log(f"   共 {len(perms):,} 种排列", 'success')
        messagebox.showinfo("完成",
                            f"✅ 表格已生成！\n\n"
                            f"排列数: {len(perms):,}\n"
                            f"文件: {excel_path}")

    # ─── 开始拼接 ───

    def _start_process(self):
        params = self._get_params()
        if not params:
            return

        groups, ordered, fixed_groups, count, res_w, res_h, output_dir, bgm_files = params
        date_str = os.path.basename(self.input_var.get())
        selected_keys = list(groups.keys())
        excel_suffix = f"固定{''.join(fixed_groups)}" if fixed_groups else None
        excel_path = get_excel_path(output_dir, date_str, ordered, selected_keys, excel_suffix)

        # 检查排列数量
        existing = load_existing_status(excel_path)
        if not existing:
            max_perms = 1
            for v in groups.values():
                max_perms *= len(v)
            if max_perms > WARN_PERM_COUNT:
                answer = messagebox.askyesno(
                    "确认",
                    f"将生成 {max_perms:,} 种排列组合，数量较大。\n\n确认继续？")
                if not answer:
                    return

        self.is_running = True
        self.stop_flag = False
        self.start_btn.config(state='disabled', bg='#9e9e9e')
        self.excel_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.result_label.config(text="")

        t = threading.Thread(target=self._run_process,
                             args=(groups, ordered, fixed_groups, count, res_w, res_h, output_dir, bgm_files),
                             daemon=True)
        t.start()

    def _stop_process(self):
        self.stop_flag = True
        self._log("⏹ 用户请求停止...", 'warn')
        self.stop_btn.config(state='disabled')

    def _run_process(self, groups, ordered, fixed_groups, count, res_w, res_h, output_dir, bgm_files=None):
        """后台线程：检查表格 → 继续未完成 或 生成全部。"""
        try:
            self.root.after(0, lambda: self._log_clear())
            self.root.after(0, lambda: self._log("🎬 开始拼接任务...", 'title'))
            mode_display = '有序' if ordered else ('无序' if not fixed_groups else f"部分固定(固定{''.join(fixed_groups)})")
            bgm_display = f" | BGM: {len(bgm_files)}首" if bgm_files else ""
            self.root.after(0, lambda md=mode_display, bd=bgm_display: self._log(
                f"模式: {md}{bd} | 本次拼接: {count} 条"))

            input_dir = self.input_var.get()
            date_str = os.path.basename(input_dir)
            selected_keys = list(groups.keys())
            excel_suffix = f"固定{''.join(fixed_groups)}" if fixed_groups else None
            excel_path = get_excel_path(output_dir, date_str, ordered, selected_keys, excel_suffix)

            # ── 检查已有表格 ──
            existing = load_existing_status(excel_path)

            if existing:
                # 表格已存在
                pending_items = {k: v for k, v in existing.items()
                                 if '✅' not in v[0]}
                done_count = len(existing) - len(pending_items)

                self.root.after(0, lambda: self._log(
                    f"📊 已有表格: {os.path.basename(excel_path)}", 'info'))
                self.root.after(0, lambda: self._log(
                    f"   ✅ 已完成: {done_count}  ⏳待拼接/❌失败: {len(pending_items)}", 'info'))

                if not pending_items:
                    self.root.after(0, lambda: self._log(
                        "🎉 全部已拼接完成！", 'success'))
                    self.root.after(0, lambda: self.result_label.config(
                        text="✅ 全部已完成", fg=self.SUCCESS))
                    return

                # 重建待拼接排列
                video_lookup = build_video_lookup(groups)
                pending_perms = []
                row_map = {}

                for key, (status, row_num, _, names) in pending_items.items():
                    # 直接用 Excel 中读取的各组视频名，不再解析排列标识
                    ordered_videos = [(video_lookup[n]['group'], video_lookup[n])
                                      for n in names if n in video_lookup]
                    if len(ordered_videos) != len(names):
                        self.root.after(0, lambda k=key: self._log(
                            f"  ⚠️ 无法解析排列 {k}，跳过", 'warn'))
                        continue

                    perm = {
                        'order': [g for g, _ in ordered_videos],
                        'videos': ordered_videos,
                        'concat_str': ' → '.join(v['name'] for _, v in ordered_videos),
                    }
                    pending_perms.append(perm)
                    row_map[perm_to_key(perm)] = row_num

                # 限制本次拼接数量
                encode_count = min(count, len(pending_perms))
                encode_perms = pending_perms[:encode_count]

                self.root.after(0, lambda: self._log(
                    f"🔄 本次拼接 {encode_count} 条（剩余 {len(pending_perms)} 条）...", 'title'))

                success, fail, skip = self._encode_loop(
                    encode_perms, groups, excel_path, output_dir, date_str,
                    row_map=row_map, target_w=res_w, target_h=res_h,
                    bgm_files=bgm_files)

            else:
                # 没有表格 → 生成全部
                self.root.after(0, lambda: self._log(
                    "未找到表格，正在生成所有排列...", 'title'))

                perms = generate_all_permutations(groups, ordered, fixed_groups)
                create_excel(perms, groups, excel_path)

                self.root.after(0, lambda: self._log(
                    f"📊 表格已创建: {os.path.basename(excel_path)} "
                    f"({len(perms):,} 种排列)", 'success'))

                # 限制本次拼接数量
                encode_count = min(count, len(perms))
                encode_perms = perms[:encode_count]

                self.root.after(0, lambda: self._log(
                    f"🔄 本次拼接 {encode_count} 条", 'title'))

                success, fail, skip = self._encode_loop(
                    encode_perms, groups, excel_path, output_dir, date_str,
                    target_w=res_w, target_h=res_h,
                    bgm_files=bgm_files)

            # ── 完成报告 ──
            result_text = (f"✅ 成功: {success}   ⏭️ 跳过: {skip}   "
                           f"❌ 失败: {fail}")
            self.root.after(0, lambda: self.result_label.config(
                text=result_text, fg=self.SUCCESS if fail == 0 else self.FAIL))
            self.root.after(0, lambda: self._log("=" * 50, 'title'))
            self.root.after(0, lambda: self._log(f"🎉 {result_text}", 'success'))
            self.root.after(0, lambda: self._log(
                f"📁 视频: {output_dir}", 'success'))
            self.root.after(0, lambda: self._log(
                f"📊 表格: {excel_path}", 'success'))

        except Exception as e:
            self.root.after(0, lambda: self._log(f"💥 严重错误: {e}", 'error'))
            self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
        finally:
            self.is_running = False
            self.root.after(0, self._reset_buttons)

    def _encode_loop(self, perms, groups, excel_path, output_dir, date_str,
                     row_map=None, target_w=None, target_h=None, bgm_files=None):
        """编码循环。row_map 为 None 时按序号计算行号。"""
        success = fail = skip = 0
        total = len(perms)

        # 初始化BGM管理器
        bgm_mgr = BgmManager(bgm_files) if bgm_files else None
        use_bgm = bgm_mgr is not None

        for idx, perm in enumerate(perms):
            if self.stop_flag:
                self.root.after(0, lambda: self._log("⏹ 已停止", 'warn'))
                break

            key = perm_to_key(perm)
            output_name = perm_to_filename(perm, date_str)
            output_path = os.path.join(output_dir, output_name)

            # Excel 行号
            if row_map:
                excel_row = row_map.get(key, idx + 2)
            else:
                excel_row = idx + 2

            # 进度
            done = success + fail + skip
            pct = (done / total * 100) if total else 0
            self.root.after(0, lambda p=pct, d=done, c=total, s=perm['concat_str']:
                self._update_progress(p, d, c, s))

            # 已存在 → 跳过
            if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                update_excel_status(excel_path, excel_row, '✅ 已完成',
                                    output_name, groups)
                skip += 1
                self.root.after(0, lambda n=output_name: self._log(
                    f"  ⏭️ {n} 已存在，跳过", 'warn'))
                continue

            # 拼接
            video_list = [vid for _, vid in perm['videos']]
            try:
                ok = concat_videos(video_list, output_path, target_w, target_h)
                if ok and use_bgm:
                    # 拼接成功，添加BGM
                    bgm = bgm_mgr.next()
                    info = probe_video(output_path)
                    bgm_output = output_path + '.bgm.mp4'
                    bgm_ok = add_bgm_to_video(output_path, bgm['path'],
                                              bgm_output, info['duration'])
                    if bgm_ok:
                        # 替换原文件
                        os.replace(bgm_output, output_path)
                        self.root.after(0, lambda n=output_name, b=bgm['name']: self._log(
                            f"  🎵 {n} + BGM: {b}", 'info'))
                    else:
                        # BGM失败不影响视频，删除临时文件
                        if os.path.exists(bgm_output):
                            os.remove(bgm_output)
                        self.root.after(0, lambda n=output_name: self._log(
                            f"  ⚠️ {n} BGM添加失败，保留原音频", 'warn'))
                if ok:
                    update_excel_status(excel_path, excel_row, '✅ 已完成',
                                        output_name, groups)
                    success += 1
                    self.root.after(0, lambda n=output_name: self._log(
                        f"  ✅ {n}", 'success'))
                else:
                    update_excel_status(excel_path, excel_row, '❌ 失败',
                                        '', groups)
                    fail += 1
                    self.root.after(0, lambda n=output_name: self._log(
                        f"  ❌ {n} 拼接失败", 'error'))
            except Exception as e:
                update_excel_status(excel_path, excel_row, '❌ 异常', '', groups)
                fail += 1
                self.root.after(0, lambda n=output_name, err=str(e): self._log(
                    f"  ❌ {n}: {err}", 'error'))

        return success, fail, skip

    def _update_progress(self, pct, done, total, current_str):
        self.progress_var.set(pct)
        self.progress_label.config(text=f"{done}/{total} ({pct:.0f}%)")
        if current_str:
            self._log(f"  [{done+1}/{total}] {current_str}")

    def _reset_buttons(self):
        self.start_btn.config(state='normal', bg=self.ACCENT)
        self.excel_btn.config(state='normal')
        self.stop_btn.config(state='disabled')


# ═══════════════════════════════════════════════════════
# 启动
# ═══════════════════════════════════════════════════════

def setup_ffmpeg_path():
    """将打包目录中的 ffmpeg 加入 PATH（PyInstaller 打包后生效）。"""
    if getattr(sys, 'frozen', False):
        # PyInstaller 打包后，ffmpeg 在 _internal/ 目录
        internal_dir = os.path.join(os.path.dirname(sys.executable), '_internal')
    else:
        # 开发环境，ffmpeg 在脚本同级目录
        internal_dir = os.path.dirname(os.path.abspath(__file__))
    if os.path.isdir(internal_dir):
        os.environ['PATH'] = internal_dir + os.pathsep + os.environ.get('PATH', '')


def check_ffmpeg():
    try:
        subprocess.run(['ffmpeg', '-version'], capture_output=True, check=True)
        return True
    except (FileNotFoundError, subprocess.CalledProcessError):
        return False


def main():
    setup_ffmpeg_path()  # 先尝试设置 ffmpeg 路径
    if not check_ffmpeg():
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "缺少依赖",
            "未检测到 ffmpeg！\n\n"
            "请安装 ffmpeg 并确保在 PATH 中：\n"
            "  winget install ffmpeg")
        return

    root = tk.Tk()
    try:
        root.iconbitmap(default='')
    except Exception:
        pass
    VideoJoinerApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
